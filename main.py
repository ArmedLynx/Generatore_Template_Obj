import openpyxl
from openpyxl.utils import get_column_letter

import csv

from os import mkdir
from os.path import exists as path_exists
from os.path import splitext as splitext

import sys, getopt


class Template():
    def __init__(self, template):
        self.template = template
        # Copio il file di testo in una stringa
        file = open(template, "rt")
        self.data = file.read()
        file.close()

    def Replace(self, tag, valore): # Sostituisce la stringa "tag" con la stringa "valore"
        self.data = self.data.replace(tag, str(valore))
    
    def Save(self, nome, out_dir="./"): # Salva il file come "nome" nel percorso "out_dir"
        outF = open(out_dir+"/"+nome, "w+")
        outF.write(self.data)
        outF.close()

class Data():
    def __init__(self, data):
        self.data = data
        # Apro il foglio di lavoro
        wb=openpyxl.load_workbook(self.data)
        self.sheet = wb.active
        # Ottengo una lista contenente le colonne utilizzate (Es.: ["A","B","C"])
        self.cols = [get_column_letter(c) for c in range(2, 1+len(tuple(self.sheet.columns)))]
    
    def GetFileName(self, riga): # Restituisce il filename contenuto nella cella alla riga "riga"
        return self.sheet["A"+str(riga)].value
    
    def GetFileNames(self): # Restitusce una stringa contenente tutti i filename ordinati
        names = [self.sheet["A"+str(i)].value for i in range(2, 1+len(tuple(self.sheet.rows)))]
        return names
    
    def GetTag(self, colonna): # Restituisce il tag contenuto nella cella alla colonna "colonna"
        return self.sheet[colonna+"1"].value
    
    def GetTags(self): # Restituisce una stringa contenente tutti i tag ordinati
        tags = [self.sheet[i+"1"].value for i in self.cols]
        return tags
    
    def GetValue(self, coord): # Restituisce il contenuto di una cella date le coordinate nel formato "A1"
        return self.sheet[coord].value
    
    def GetCell(self, file, tag): # Restituisce le coordinate della cella con colonna corrispondente a quella del tag e riga corrispondente a quella del filename
        col = self.cols[self.GetTags().index(tag)]
        row = self.GetFileNames().index(file)+2
        return col+str(row)
        

class CsvData():
    def __init__(self, data):
        self.data=data
        #Apro il file CSV e lo salvo in una matrice
        self.csvMatrix=[]
        with open(self.data, 'r') as file:
           reader=csv.reader(file, delimiter=';')
           for row in reader:
             self.csvMatrix.append(row)
    
    def GetFileName(self, riga): # restituisce il nome file alla riga indicata
        return self.csvMatrix[riga][0]
    
    def GetFileNames(self): # Restituisce una stringa contenente tutti i nomi file
        names = [riga[0] for riga in self.csvMatrix]
        names.pop(0) # Elimino il primo elemento della lista perchè è "f_name"
        return names

    def GetTag(self, colonna): # Restituisce il tag alla colonna indicata
        return self.csvMatrix[0][colonna]

    def GetTags(self): # Restituisce una lista contente tutti i tag
        tags=list(self.csvMatrix[0])
        tags.pop(0) # Elimino il primo elemento della lista perchè è "f_name"
        return tags

    def GetValue(self, coord): # Restituisce il contenuto di una cella date le coordinate nel formato "A1"
        col=ord(coord[0])-65 # converte il char nell'int corrispondente al suo codice ascii
        row=int(coord[1])-1
        return self.csvMatrix[row][col]

    def GetCell(self, file, tag): # Restituisce le coordinate della cella con colonna corrispondente a quella del tag e riga corrispondente a quella del filename
        col=self.GetTags().index(tag)+1
        row=self.GetFileNames().index(file)+2
        return str(chr(col+65))+str(row)
    
    # def PrintTest(self):
    #     print(self.csvMatrix)



def main(argv):
    opts, args = getopt.getopt(argv, "t:d:o:Fh") # Eseguo il parsing della lista di argomenti per ottenere una lista di tuple (opzione, argomento)

    aiuto = '''
        -t <Path>\tPercorso per il file template
        -d <Path>\tPercorso per il file data
        -o <Path>\tPercorso per la cartella di output
        -F\tForza la riscrittura dei file di configurazione già creati
        -h\tAiuto
        '''
    # Assegno alle variabili il valore di default
    template_file = "./template.txt"
    data_file = "./data.csv"
    out_path = "./outputs"
    overwrite = False

    # Assegno alle variabili il valore passato da cmd
    for opt, arg in opts:
        if opt == "-t":
            template_file = arg
        elif opt == "-d":
            data_file = arg
        elif opt == "-o":
            out_path = arg
        elif opt == "-F":
            overwrite = True
        elif opt == "-h":
            print(aiuto)
            return
        else:
            print("Parametri non validi.\n"+aiuto)

    # Se la cartella di output non esiste la creo
    if path_exists(out_path) == False:
        print(": Creo la cartella di output\n|")
        mkdir(out_path)
    
    # Recupero l'estensione del file data
    print(": Verifico il formato del file data")
    file_name, file_extension = splitext(data_file)
    if file_extension == ".csv" or file_extension == ".CSV":
        #Creo un oggetto contenente il file csv
        print(":- Leggo il file csv\n|")
        xls = CsvData(data_file)
    elif file_extension == ".xlsx" or file_extension == ".XLSX":
        # Creo un oggetto contenente il folgio di lavoro xlsx
        print(":- Leggo il file xlsx\n|")
        xls = Data(data_file)
    else:
        print("# ERR: Formato file data non supportato")
        return
    
    # Eseguo un ciclo sulla lista contente i filename
    for file in xls.GetFileNames():
        # Se il flag overwrite è False e il file esiste non faccio nulla altrimenti creo il file configurazione
        if overwrite == False and path_exists(out_path+"/"+file+".txt") == True:
            print(": Il file "+file+".txt è già presente")
        else:
            # Creo un oggetto contenente il template
            conf = Template(template_file)
            print(":- Apro il template per "+file)
            # Eseguo un ciclo sulla lista contenete i tags
            for tag in xls.GetTags():
                # Sostituisco il tag con il contenuto della cella alle coordinate corrispondenti al file e al tag
                conf.Replace(tag, xls.GetValue(xls.GetCell(file, tag)))
                print(":-- Sostituisco "+tag+" con "+str(xls.GetValue(xls.GetCell(file, tag))))
            # Salvo il file di configurazione
            conf.Save(file+".txt", out_path)
            print(":- Salvo il file di cofigurazione "+file+"\n|")
    
    print(": Ho terminato la preparazione delle configurazioni")
    
# def test():
#     xls = CsvData(r'C:\Users\lspreafico.MATICMINDIT\Desktop\Cartel1.csv')
#     print(xls.GetFileNames())
#     for i in range (1, 3):
#         print(xls.GetFileName(i))
#     print(xls.GetTags())
#     for i in range (1, 4):
#         print(xls.GetTag(i))
#     print(xls.GetValue("B2"))
#     print(xls.GetValue("C3"))
#     print(xls.GetCell("pippo", "<NET>"))
#     print(xls.GetCell("pluto", "<HOST>"))
#     print(xls.GetValue(xls.GetCell("pippo", "<NET>")))
#     print(xls.GetValue(xls.GetCell("pluto", "<HOST>")))

if __name__ == "__main__":
    main(sys.argv[1:])
    # test()