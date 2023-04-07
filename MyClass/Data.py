import openpyxl
from openpyxl.utils import get_column_letter

import csv

class Data():
    def __init__(self, data):
        self.data = data
        # Apro il foglio di lavoro
        wb=openpyxl.load_workbook(self.data)
        self.sheet = wb.active
        # Ottengo una lista contenente le colonne utilizzate (Es.: ["A","B","C"])
        self.cols = [get_column_letter(c) for c in range(2, 1+len(tuple(self.sheet.columns)))]
    
    def GetFileName(self, riga): # Restituisce il filename contenuto nella cella alla riga "riga"
        return self.sheet["A"+str(riga)].value
    
    def GetFileNames(self): # Restitusce una stringa contenente tutti i filename ordinati
        names = [self.sheet["A"+str(i)].value for i in range(2, 1+len(tuple(self.sheet.rows)))]
        return names
    
    def GetTag(self, colonna): # Restituisce il tag contenuto nella cella alla colonna "colonna"
        return self.sheet[colonna+"1"].value
    
    def GetTags(self): # Restituisce una stringa contenente tutti i tag ordinati
        tags = [self.sheet[i+"1"].value for i in self.cols]
        return tags
    
    def GetValue(self, coord): # Restituisce il contenuto di una cella date le coordinate nel formato "A1"
        return self.sheet[coord].value
    
    def GetCell(self, file, tag): # Restituisce le coordinate della cella con colonna corrispondente a quella del tag e riga corrispondente a quella del filename
        col = self.cols[self.GetTags().index(tag)]
        row = self.GetFileNames().index(file)+2
        return col+str(row)
        

class CsvData():
    def __init__(self, data):
        self.data=data
        #Apro il file CSV e lo salvo in una matrice
        self.csvMatrix=[]
        with open(self.data, 'r') as file:
           reader=csv.reader(file, delimiter=';')
           for row in reader:
             self.csvMatrix.append(row)
    
    def GetFileName(self, riga): # restituisce il nome file alla riga indicata
        return self.csvMatrix[riga][0]
    
    def GetFileNames(self): # Restituisce una stringa contenente tutti i nomi file
        names = [riga[0] for riga in self.csvMatrix]
        names.pop(0) # Elimino il primo elemento della lista perchè è "f_name"
        return names

    def GetTag(self, colonna): # Restituisce il tag alla colonna indicata
        return self.csvMatrix[0][colonna]

    def GetTags(self): # Restituisce una lista contente tutti i tag
        tags=list(self.csvMatrix[0])
        tags.pop(0) # Elimino il primo elemento della lista perchè è "f_name"
        return tags

    def GetValue(self, coord): # Restituisce il contenuto di una cella date le coordinate nel formato "A1"
        col=ord(coord[0])-65 # converte il char nell'int corrispondente al suo codice ascii
        row=int(coord[1])-1
        return self.csvMatrix[row][col]

    def GetCell(self, file, tag): # Restituisce le coordinate della cella con colonna corrispondente a quella del tag e riga corrispondente a quella del filename
        col=self.GetTags().index(tag)+1
        row=self.GetFileNames().index(file)+2
        return str(chr(col+65))+str(row)
    
    # def PrintTest(self):
    #     print(self.csvMatrix)